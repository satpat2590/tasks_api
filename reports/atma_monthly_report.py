#!/usr/bin/env python3
"""Atma Monthly Report Generator — creates Excel reports and uploads to Google Drive.

Generates a per-month Excel workbook with:
  - Sheet 1: All task completions (task, category, user, points, quality, date)
  - Sheet 2: Category breakdown (completions, total points, avg quality)
  - Sheet 3: Per-user summary (tasks completed, points earned/lost)

Usage:
    python -m reports.atma_monthly_report --month 2026-06
    python -m reports.atma_monthly_report --month 2026-06 --drive
"""

import argparse
import logging
import os
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Optional

import requests

logger = logging.getLogger(__name__)

ATMA_BASE_URL = os.getenv("ATMA_BASE_URL", "https://tasks-api-71v5.onrender.com")
ATMA_TOKEN = os.getenv("ATMA_BEARER_TOKEN", os.getenv("ATMA_HERMES_TOKEN", ""))


def fetch_completions(limit: int = 500) -> List[dict]:
    """Fetch all task completions from Atma API, paginating to avoid truncation."""
    headers = {"Authorization": f"Bearer {ATMA_TOKEN}"}
    page_size = 500
    completions: List[dict] = []
    offset = 0
    while True:
        r = requests.get(
            f"{ATMA_BASE_URL}/api/agent/completed",
            headers=headers,
            params={"limit": page_size, "offset": offset},
            timeout=30,
        )
        r.raise_for_status()
        batch = r.json()
        if not batch:
            break
        completions.extend(batch)
        if len(batch) < page_size:
            break
        offset += len(batch)
        if len(completions) >= limit:
            break
    return completions[:limit]


def fetch_active_tasks() -> List[dict]:
    """Fetch all active tasks for the user summary."""
    headers = {"Authorization": f"Bearer {ATMA_TOKEN}"}
    r = requests.get(
        f"{ATMA_BASE_URL}/api/agent/tasks",
        headers=headers,
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def filter_by_month(completions: List[dict], month: str) -> List[dict]:
    """Filter completions to a specific YYYY-MM month."""
    return [c for c in completions if c.get("completed_at", "").startswith(month)]


def generate_excel_report(completions: List[dict], active_tasks: List[dict], month: str) -> str:
    """Generate an Excel report and return the file path.

    Uses openpyxl to create a formatted .xlsx workbook with:
    - Sheet 1: All completions
    - Sheet 2: Category breakdown
    - Sheet 3: Active tasks at end of month
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────────────
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    positive_font = Font(color="008000")
    negative_font = Font(color="CC0000")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)

    # ── Sheet 1: All Completions ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Completions"

    ws1["A1"] = f"Atma Monthly Report — {month}"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:G1")

    headers = ["Date", "Task ID", "Task Title", "Category", "Points", "Quality", "Late?"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=h)
    style_header(ws1, 3)

    for i, c in enumerate(sorted(completions, key=lambda x: x.get("completed_at", "")), start=4):
        ws1.cell(row=i, column=1, value=c.get("completed_at", "")[:10])
        ws1.cell(row=i, column=2, value=c.get("task_id"))
        ws1.cell(row=i, column=3, value=c.get("task_title", ""))
        ws1.cell(row=i, column=4, value=c.get("task_category", ""))
        pts_cell = ws1.cell(row=i, column=5, value=c.get("points", 0))
        pts_cell.font = positive_font if (c.get("points", 0) >= 0) else negative_font
        ws1.cell(row=i, column=6, value=c.get("quality", "") or "")
        ws1.cell(row=i, column=7, value="Yes" if c.get("was_late") else "No")

    auto_width(ws1)

    # ── Sheet 2: Category Breakdown ─────────────────────────────────
    ws2 = wb.create_sheet("Category Breakdown")

    ws2["A1"] = f"Category Breakdown — {month}"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:F1")

    cat_headers = ["Category", "Completions", "Total Points", "Avg Points", "Late Count", "Avg Quality"]
    for col, h in enumerate(cat_headers, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header(ws2, 3)

    cat_data = defaultdict(lambda: {"count": 0, "points": 0, "late": 0, "quality_sum": 0, "quality_count": 0})
    for c in completions:
        cat = c.get("task_category", "unknown")
        cat_data[cat]["count"] += 1
        cat_data[cat]["points"] += c.get("points", 0)
        if c.get("was_late"):
            cat_data[cat]["late"] += 1
        q = c.get("quality")
        if q and isinstance(q, (int, float)):
            cat_data[cat]["quality_sum"] += q
            cat_data[cat]["quality_count"] += 1

    row = 4
    for cat in sorted(cat_data.keys()):
        d = cat_data[cat]
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=d["count"])
        pts_cell = ws2.cell(row=row, column=3, value=d["points"])
        pts_cell.font = positive_font if d["points"] >= 0 else negative_font
        ws2.cell(row=row, column=4, value=round(d["points"] / d["count"], 1) if d["count"] else 0)
        ws2.cell(row=row, column=5, value=d["late"])
        ws2.cell(row=row, column=6, value=round(d["quality_sum"] / d["quality_count"], 1) if d["quality_count"] else "N/A")
        row += 1

    # Totals row
    total_count = sum(d["count"] for d in cat_data.values())
    total_pts = sum(d["points"] for d in cat_data.values())
    total_late = sum(d["late"] for d in cat_data.values())
    ws2.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=total_count).font = Font(bold=True)
    pts_cell = ws2.cell(row=row, column=3, value=total_pts)
    pts_cell.font = Font(bold=True, color="008000" if total_pts >= 0 else "CC0000")
    ws2.cell(row=row, column=4, value=round(total_pts / total_count, 1) if total_count else 0).font = Font(bold=True)
    ws2.cell(row=row, column=5, value=total_late).font = Font(bold=True)

    auto_width(ws2)

    # ── Sheet 3: Active Tasks ───────────────────────────────────────
    ws3 = wb.create_sheet("Active Tasks")

    ws3["A1"] = f"Active Tasks (as of report date) — {month}"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:E1")

    task_headers = ["ID", "Title", "Category", "Priority", "Recurring?"]
    for col, h in enumerate(task_headers, 1):
        ws3.cell(row=3, column=col, value=h)
    style_header(ws3, 3)

    for i, t in enumerate(sorted(active_tasks, key=lambda x: x.get("id", 0)), start=4):
        ws3.cell(row=i, column=1, value=t.get("id"))
        ws3.cell(row=i, column=2, value=t.get("title", ""))
        ws3.cell(row=i, column=3, value=t.get("category", ""))
        ws3.cell(row=i, column=4, value=t.get("priority", ""))
        ws3.cell(row=i, column=5, value="Yes" if t.get("is_recurring") else "No")

    auto_width(ws3)

    # ── Save ────────────────────────────────────────────────────────
    tmpdir = tempfile.mkdtemp()
    filename = f"atma_monthly_report_{month}.xlsx"
    filepath = os.path.join(tmpdir, filename)
    wb.save(filepath)
    logger.info(f"Report saved: {filepath}")
    return filepath


def upload_to_drive(filepath: str, month: str) -> str:
    """Upload report to Google Drive under Omni/Atma/reports/."""
    # Try omni-drive package first, fall back to Edoras drive_uploader
    try:
        from omni_drive import DriveClient

        client = DriveClient()
        filename = os.path.basename(filepath)
        link = client.upload("Atma", filename, filepath, subfolder="reports")
        logger.info(f"Uploaded to Omni/Atma/reports/: {link}")
        return link
    except ImportError:
        pass

    # Fallback: Edoras drive_uploader
    try:
        sys.path.insert(0, os.path.expanduser("~/edoras/src"))
        from reports.drive_uploader import DriveUploader

        uploader = DriveUploader()
        filename = os.path.basename(filepath)
        link = uploader.upload(filename, f"Atma/reports/{filename}")
        logger.info(f"Uploaded via Edoras uploader: {link}")
        return link
    except Exception as e:
        logger.error(f"Drive upload failed: {e}")
        return None


def main():
    parser = argparse.ArgumentParser(description="Atma Monthly Report Generator")
    parser.add_argument("--month", required=True, help="Month in YYYY-MM format (e.g., 2026-06)")
    parser.add_argument("--drive", action="store_true", help="Upload to Google Drive after generating")
    parser.add_argument("--output", help="Custom output path (default: temp dir)")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    logger.info(f"Generating Atma monthly report for {args.month}")

    # Fetch data
    completions = fetch_completions()
    month_completions = filter_by_month(completions, args.month)
    logger.info(f"Found {len(month_completions)} completions for {args.month}")

    active_tasks = fetch_active_tasks()
    logger.info(f"Found {len(active_tasks)} active tasks")

    if not month_completions:
        logger.warning(f"No completions found for {args.month}")

    # Generate Excel
    filepath = generate_excel_report(month_completions, active_tasks, args.month)

    if args.output:
        import shutil

        shutil.copy(filepath, args.output)
        filepath = args.output

    print(f"\nReport: {filepath}")

    # Upload to Drive
    if args.drive:
        link = upload_to_drive(filepath, args.month)
        if link:
            print(f"Drive: {link}")
        else:
            print("Drive upload failed — report saved locally")


if __name__ == "__main__":
    main()
